import requests
from dotenv import load_dotenv
import os
import googlemaps
import pandas as pd
import openpyxl

load_dotenv()

api_key = os.getenv('API_KEY')

def get_mandal():
  df = pd.read_excel("schools.xlsx")
  mandals = df["MANDAL"].tolist()
  return mandals

def get_school():
  df = pd.read_excel("schools.xlsx")
  schools = df["School Name"].tolist()
  return schools

# demo

# Function to insert data into Excel sheet
def insert_data_into_excel(iterable_object, excel_file, column_name, column_number):
    try:
        # Read the existing Excel file into a DataFrame
        df = pd.read_excel(excel_file)
        
        # Check if the specified column name already exists, if not, create it
        if column_name not in df.columns:
            df[column_name] = None
        
        # Iterate through the iterable object and insert data into the specified column
        for i, data in enumerate(iterable_object):
            df.at[i, column_name] = data
        
        # Write the updated DataFrame back to the Excel file
        df.to_excel(excel_file, index=False)
        print(f"Data inserted into '{column_name}' in '{excel_file}' successfully.")
    except Exception as e:
        print(f"An error occurred: {str(e)}")


  
def get_distance(origin, destination):
    client = googlemaps.Client(key=api_key)
    try:
      distance_matrix = client.distance_matrix(origins=[origin], destinations=[destination])
      dis = distance_matrix["rows"][0]["elements"][0]["distance"]["value"]
      dis = dis/1000
      return dis
    except KeyError:
      return "Not Found"

if __name__ == "__main__":
  origin = "18.43455403564981, 79.10711311605102"
  schools = get_school()
  mandals = get_mandal()
  
  destination = list()
  distances = list()

  
  insert_data_into_excel(mandals, 'output1.xlsx', "Mandal Name", 2)
  insert_data_into_excel(schools, 'output1.xlsx', "School Name", 4)
  

  for i in range(len(mandals)):
    if (mandals[i] == "CHIGURUMAMIDI"):
      destination.append(f"{schools[i]},Chigurumamidi")
  for i in range(len(mandals)):
    if (mandals[i] == "CHOPPADANDI"):
      destination.append(f"{schools[i]},Choppadandi")
  for i in range(len(mandals)):
    if (mandals[i] == "ELLANDAKUNTA"):
      destination.append(f"{schools[i]},Ellandakunta")
  for i in range(len(mandals)):
    if (mandals[i] == "GANGADHARA"):
      destination.append(f"{schools[i]},Gangadhara")
  for i in range(len(mandals)):
    if (mandals[i] == "GANNERVARAM"):
      destination.append(f"{schools[i]},Gannervaram")
  for i in range(len(mandals)):
    if (mandals[i] == "HUZURABAD"):
      destination.append(f"{schools[i]},Huzurabad")
  for i in range(len(mandals)):
    if (mandals[i] == "KARIMNAGAR"):
      destination.append(f"{schools[i]},Karimnagar")
  for i in range(len(mandals)):
    if (mandals[i] == "KARIMNAGAR (RURAL)"):
      destination.append(f"{schools[i]},Karimnagar (Rural)")
  for i in range(len(mandals)):
    if (mandals[i] == "KOTHAPALLE"):
      destination.append(f"{schools[i]},Kothapalle")
  for i in range(len(mandals)):
    if (mandals[i] == "MANAKONDUR"):
      destination.append(f"{schools[i]},Manakondur")
  for i in range(len(mandals)):
    if (mandals[i] == "RAMADUGU"):
      destination.append(f"{schools[i]},Ramadugu")
  for i in range(len(mandals)):
    if (mandals[i] == "SAIDAPUR (V)"):
      destination.append(f"{schools[i]},Saipadur")
  for i in range(len(mandals)):
    if (mandals[i] == "SHANKARAPATNAM"):
      destination.append(f"{schools[i]},Shankarapatnam")
  for i in range(len(mandals)):
    if (mandals[i] == "THIMMAPUR"):
      destination.append(f"{schools[i]},Thimmapur")
  for i in range(len(mandals)):
    if (mandals[i] == "VEENAVANKA"):
      destination.append(f"{schools[i]},Veenavanka")
    
  for i in range(len(destination)):
      distances.append(get_distance(origin, destination[i]))
      
  print(len(distances))
  
  insert_data_into_excel(distances, 'output1.xlsx', "Distance", 6)
      
  
# ----------------------------------------------
    
   # distances[i] = round(distances[i],3)

  # mandals = extract_rows_and_insert_value("MANDAL", "THIMMAPUR", 0)
  
  # print(len(schools))
  # print(len(mandals))
# def extract_rows_and_insert_value( column_name, filter_value, start_value):

#   df = pd.read_excel("schools.xlsx")
#   filtered_df = df[df["MANDAL"] == filter_value]

#   # Get the rows from the specified value to the end of the DataFrame.
#   rows_to_extract = filtered_df[filtered_df[column_name] == start_value:]

#   # Insert the value into the list.
#   # rows_to_extract[column_name] = insert_value

#   # Return the list of column contents from the filtered rows.
#   return rows_to_extract[column_name].tolist()



  # print(destination)
#   print(f"The distance between {origin} and {destination} is {distance/1000} kilometers.")




# def create_excel_sheet(iterable_object_data, column_name,column_no, wb=None):
#   wb = openpyxl.Workbook()

#   ws = wb.create_sheet()

#   ws.cell(row=1, column=column_no).value = column_name

#   for row_index, row_data in enumerate(iterable_object_data):
#     ws.cell(row=row_index + 2, column=column_no).value = row_data

#   wb.save('output.xlsx')